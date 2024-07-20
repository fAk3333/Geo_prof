import pandas as pd


def txt_to_ary(file_name):
    with open(file_name, encoding="utf-8") as file:
        array = [row.strip() for row in file]

    new_arr = []
    for string in array:
        new_arr.append(string.split())

    piketi = []
    otmetki = []
    nazvaniya = []
    ugod = []

    for ary in new_arr:
        if len(ary) != 4:
            ary.append('')

    for i in range(len(new_arr)):
        for j in range(len(new_arr)):
            if j == 0:
                piketi.append(new_arr[i][j])
            elif j == 1:
                otmetki.append(new_arr[i][j])
            elif j == 2:
                nazvaniya.append(new_arr[i][j])
            elif j == 3:
                if new_arr[i][j] != '':
                    ugod.append(new_arr[i][j])
                else:
                    ugod.append(new_arr[i-1][j])
    for i in range(len(ugod)):
        if ugod[i] == '':
            ugod[i] = ugod[i-1]
    return piketi, otmetki, nazvaniya, ugod
# print(txt_to_ary('tst.txt'))


def scaling_otmetok(otmetki):
    tmp_ary = [float(i) for i in otmetki]
    min_otm = min(tmp_ary)
    new_otmetki = []
    for i in tmp_ary:
        if i == min:
            new_otmetki.append(60)
        else:
            new_otmetki.append(((i - min_otm) * 10 + 60).__round__(2))

    return new_otmetki
# print(scaling_otmetok(['123.15', '123.20', '123.50', '123.40']))


def scaling_piketov(piketi):
    #pk * 2 input format: '0+00.0'
    new_piketi = []
    for i in piketi:
        big_num = int(i[0])
        small_num = float(i[2:])
        res = (big_num*100 + small_num)*2
        new_piketi.append(res)
    return new_piketi
#print(scaling_piketov(['0+00.0', '0+05.4', '0+11.2', '0+13.4', '1+05.0', '2+50.0']))

def xlsx_to_ary(file_name):
    from openpyxl import load_workbook
    wb = load_workbook(file_name)
    ws = wb.active
    column = ws['A']
    PK_lst = [column[x].value for x in range(len(column))]
    for elem in PK_lst:
        if elem == ' ' or elem == None or elem is None:
            PK_lst.remove(elem)
    PK_lst = [x for x in PK_lst if x is not None]
    column = ws['B']
    columnU = ws['D']
    ugods = [columnU[x].value for x in range(len(column))]
    otmetki = [column[x].value for x in range(len(column))]
    for i in range(len(otmetki)):
        for j in range(len(ugods)):
            if otmetki[i] != None and ugods[j] == None:
                ugods[j] = ugods[j-1]
    otmetki, ugods = zip(*[(x, y) for x, y in zip(otmetki, ugods) if x is not None])
    otmetki = list(otmetki)
    otmetki = list(map(str, otmetki))
    column = ws['C']
    names = [column[x].value for x in range(len(column))]
    for elem in names:
        if elem == None or elem is None:
            names.remove(elem)
    names = [x for x in names if x is not None]
    names = [s.replace(' ', '') for s in names]
    ugods = list(ugods)
    return PK_lst, otmetki, names, ugods
